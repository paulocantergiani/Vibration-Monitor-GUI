"""
Módulo de exportação de relatórios de monitoramento de vibração.
Instituto Militar de Engenharia
Projeto: Monitoramento Inteligente de Carga

Suporta exportação em PDF e XLSX com gráficos e estatísticas.
"""

import os
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from io import BytesIO

import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Backend não-interativo


class ReportExporter:
    """Classe base para exportação de relatórios."""

    def __init__(self, sensor_id: str, unit: str = "ADC"):
        """
        Inicializa o exportador.

        Args:
            sensor_id: Identificador do sensor
            unit: Unidade de medida (padrão: ADC)
        """
        self.sensor_id = sensor_id
        self.unit = unit
        self.timestamp = datetime.now()

    def _prepare_data(self, data_list: List[Dict]) -> pd.DataFrame:
        """Converte lista de dados em DataFrame."""
        df = pd.DataFrame(data_list)
        if 'timestamp' in df.columns:
            df['timestamp'] = pd.to_datetime(df['timestamp'])
        return df


class PDFReportExporter(ReportExporter):
    """Exportador de relatórios em formato PDF."""

    def export(self,
               filename: str,
               data_list: List[Dict],
               stats: Dict,
               graph_image_path: Optional[str] = None) -> bool:
        """
        Exporta relatório em PDF.

        Args:
            filename: Caminho do arquivo de saída
            data_list: Lista de dados de vibração
            stats: Dicionário com estatísticas
            graph_image_path: Caminho para imagem do gráfico

        Returns:
            True se sucesso, False caso contrário
        """
        try:
            # Prepara dados
            df = self._prepare_data(data_list)

            # Cria documento PDF
            doc = SimpleDocTemplate(
                filename,
                pagesize=letter,
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=0.75*inch,
                bottomMargin=0.75*inch
            )

            elements = []
            styles = getSampleStyleSheet()

            # Título do relatório
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1f77b4'),
                spaceAfter=6,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )

            elements.append(Paragraph("RELATÓRIO DE MONITORAMENTO", title_style))
            elements.append(Paragraph("Dados de Vibração", styles['Heading2']))
            elements.append(Spacer(1, 0.3*inch))

            # Informações gerais
            info_data = [
                ["Sensor ID", self.sensor_id],
                ["Data/Hora", self.timestamp.strftime("%d/%m/%Y %H:%M:%S")],
                ["Unidade", self.unit],
                ["Total de Leituras", str(len(df))],
            ]

            info_table = Table(info_data, colWidths=[2.5*inch, 3.5*inch])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8f4f8')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))

            elements.append(info_table)
            elements.append(Spacer(1, 0.3*inch))

            # Estatísticas
            elements.append(Paragraph("Estatísticas", styles['Heading2']))
            elements.append(Spacer(1, 0.15*inch))

            stats_data = [
                ["Métrica", "Valor"],
                ["Mínimo", f"{stats.get('min', 0):.2f} {self.unit}"],
                ["Máximo", f"{stats.get('max', 0):.2f} {self.unit}"],
                ["Média", f"{stats.get('avg', 0):.2f} {self.unit}"],
                ["Eventos de Alerta", str(stats.get('alerts', 0))],
            ]

            stats_table = Table(stats_data, colWidths=[2.5*inch, 3.5*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ]))

            elements.append(stats_table)
            elements.append(Spacer(1, 0.3*inch))

            # Gráfico se disponível
            if graph_image_path and os.path.exists(graph_image_path):
                elements.append(Paragraph("Gráfico Histórico", styles['Heading2']))
                elements.append(Spacer(1, 0.15*inch))
                try:
                    img = Image(graph_image_path, width=6*inch, height=3*inch)
                    elements.append(img)
                    elements.append(Spacer(1, 0.2*inch))
                except Exception as e:
                    print(f"Aviso: Não foi possível incluir gráfico: {e}")

            # Tabela de dados (resumo das últimas 20 leituras)
            elements.append(PageBreak())
            elements.append(Paragraph("Dados Detalhados (Últimas 20 Leituras)", styles['Heading2']))
            elements.append(Spacer(1, 0.15*inch))

            # Pega últimas 20 leituras
            df_last = df.tail(20).reset_index(drop=True)

            table_data = [["#", "Timestamp", "Valor", "Unidade"]]
            for idx, row in df_last.iterrows():
                timestamp_str = row['timestamp'].strftime("%H:%M:%S") if hasattr(row['timestamp'], 'strftime') else str(row['timestamp'])
                table_data.append([
                    str(idx + 1),
                    timestamp_str,
                    f"{row['value']:.2f}",
                    row.get('unit', self.unit)
                ])

            data_table = Table(table_data, colWidths=[0.8*inch, 2.5*inch, 1.5*inch, 1.2*inch])
            data_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ]))

            elements.append(data_table)

            # Rodapé
            elements.append(Spacer(1, 0.5*inch))
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.grey,
                alignment=TA_CENTER
            )
            elements.append(Paragraph(
                f"Relatório gerado em {self.timestamp.strftime('%d/%m/%Y às %H:%M:%S')}",
                footer_style
            ))

            # Constrói PDF
            doc.build(elements)
            return True

        except Exception as e:
            print(f"Erro ao exportar PDF: {e}")
            return False


class XLSXReportExporter(ReportExporter):
    """Exportador de relatórios em formato XLSX."""

    def export(self,
               filename: str,
               data_list: List[Dict],
               stats: Dict,
               graph_image_path: Optional[str] = None) -> bool:
        """
        Exporta relatório em XLSX.

        Args:
            filename: Caminho do arquivo de saída
            data_list: Lista de dados de vibração
            stats: Dicionário com estatísticas
            graph_image_path: Caminho para imagem do gráfico

        Returns:
            True se sucesso, False caso contrário
        """
        try:
            # Prepara dados
            df = self._prepare_data(data_list)

            # Cria workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório"

            # Define estilos
            title_font = Font(size=18, bold=True, color="FFFFFF")
            title_fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
            header_font = Font(size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
            section_font = Font(size=12, bold=True, color="1f77b4")
            data_fill = PatternFill(start_color="e8f4f8", end_color="e8f4f8", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            row = 1

            # Título
            ws.merge_cells(f'A{row}:D{row}')
            title_cell = ws[f'A{row}']
            title_cell.value = "RELATÓRIO DE MONITORAMENTO DE VIBRAÇÃO"
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row].height = 25
            row += 2

            # Informações gerais
            ws[f'A{row}'].value = "Sensor ID"
            ws[f'B{row}'].value = self.sensor_id
            row += 1

            ws[f'A{row}'].value = "Data/Hora"
            ws[f'B{row}'].value = self.timestamp.strftime("%d/%m/%Y %H:%M:%S")
            row += 1

            ws[f'A{row}'].value = "Unidade"
            ws[f'B{row}'].value = self.unit
            row += 1

            ws[f'A{row}'].value = "Total de Leituras"
            ws[f'B{row}'].value = len(df)
            row += 2

            # Seção de estatísticas
            ws[f'A{row}'].value = "ESTATÍSTICAS"
            ws[f'A{row}'].font = section_font
            row += 1

            ws[f'A{row}'].value = "Mínimo"
            ws[f'B{row}'].value = stats.get('min', 0)
            ws[f'C{row}'].value = self.unit
            row += 1

            ws[f'A{row}'].value = "Máximo"
            ws[f'B{row}'].value = stats.get('max', 0)
            ws[f'C{row}'].value = self.unit
            row += 1

            ws[f'A{row}'].value = "Média"
            ws[f'B{row}'].value = stats.get('avg', 0)
            ws[f'C{row}'].value = self.unit
            row += 1

            ws[f'A{row}'].value = "Eventos de Alerta"
            ws[f'B{row}'].value = stats.get('alerts', 0)
            row += 2

            # Seção de dados
            ws[f'A{row}'].value = "DADOS DETALHADOS"
            ws[f'A{row}'].font = section_font
            row += 1

            # Cabeçalhos da tabela
            headers = ['#', 'Timestamp', 'Valor', 'Unidade']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            row += 1

            # Dados
            for idx, (_, record) in enumerate(df.iterrows(), 1):
                ws.cell(row=row, column=1).value = idx
                timestamp_str = record['timestamp'].strftime("%d/%m/%Y %H:%M:%S") if hasattr(record['timestamp'], 'strftime') else str(record['timestamp'])
                ws.cell(row=row, column=2).value = timestamp_str
                ws.cell(row=row, column=3).value = record['value']
                ws.cell(row=row, column=4).value = record.get('unit', self.unit)

                # Aplica formatação
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                    if idx % 2 == 0:
                        cell.fill = data_fill

                row += 1

            # Adiciona gráfico se disponível
            if graph_image_path and os.path.exists(graph_image_path):
                row += 1
                ws.merge_cells(f'A{row}:D{row}')
                ws[f'A{row}'].value = "GRÁFICO HISTÓRICO"
                ws[f'A{row}'].font = section_font
                row += 1

                try:
                    img = XLImage(graph_image_path)
                    img.width = 500
                    img.height = 300
                    ws.add_image(img, f'A{row}')
                    row += 15  # Espaço para a imagem
                except Exception as e:
                    print(f"Aviso: Não foi possível incluir gráfico: {e}")

            # Ajusta largura das colunas
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12

            # Salva arquivo
            wb.save(filename)
            return True

        except Exception as e:
            print(f"Erro ao exportar XLSX: {e}")
            return False


def export_report(export_format: str,
                 filename: str,
                 sensor_id: str,
                 data_list: List[Dict],
                 stats: Dict,
                 unit: str = "ADC",
                 graph_image_path: Optional[str] = None) -> bool:
    """
    Função auxiliar para exportar relatório.

    Args:
        export_format: Formato desejado ('pdf' ou 'xlsx')
        filename: Caminho do arquivo de saída
        sensor_id: ID do sensor
        data_list: Lista de dados
        stats: Dicionário com estatísticas
        unit: Unidade de medida
        graph_image_path: Caminho para imagem do gráfico

    Returns:
        True se sucesso, False caso contrário
    """
    if export_format.lower() == 'pdf':
        exporter = PDFReportExporter(sensor_id, unit)
    elif export_format.lower() == 'xlsx':
        exporter = XLSXReportExporter(sensor_id, unit)
    else:
        raise ValueError(f"Formato não suportado: {export_format}")

    return exporter.export(filename, data_list, stats, graph_image_path)
